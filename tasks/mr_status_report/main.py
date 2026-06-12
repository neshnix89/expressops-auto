#!/usr/bin/env python3
"""
MR Status Report — Pilot Run & DMR Manufacturing Readiness
==========================================================
Migrated into the expressops-auto framework from the standalone
`Pilot_DMR_Report.py`. Behaviour is preserved; the only functional changes are:

  * Credentials/URLs/page-id are loaded from config.yaml via core.config_loader
    (no more hard-coded PATs — safe to keep in git).
  * --mock / --live / --dry-run CLI flags.
  * NEW "Status" tick-box column on the Confluence Active MR table. Ticking a
    box marks that container done on the NEXT run (it moves to COMPLETED MR) —
    a manual "settle" path for projects that do not need to go for MR, in
    addition to the existing PRSG-released and MR-Status=DONE rules.

Pulls NPI container data from Jira, classifies as Pilot Run or DMR,
cross-references with EDM Oracle DB for PRSG release status, and publishes to:
  1. Confluence page (live team view, editable manual fields)
  2. Excel file (local backup)

IMPORTANT: For the EDM/PRSG lookup to work the live run must use EDMAdmin.exe
(a renamed python.exe) to bypass SYS.PF_SEC_LOGON_TRIGGER:
  EDMAdmin.exe -m tasks.mr_status_report.main --live
Under a plain python.exe the EDM step is skipped gracefully (no PRSG statuses).
"""
from __future__ import annotations

import argparse
import re
import sys
import warnings
import logging
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
import urllib3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

from core.config_loader import load_config  # noqa: E402

urllib3.disable_warnings()
warnings.filterwarnings('ignore')

# =====================================================================
# CONFIGURATION
# =====================================================================
# --- Logging ---
LOG_DIR = ROOT / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOG_DIR / "mr_status_report.log"
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()  # Also print to console
    ]
)
log = logging.getLogger("MR_Report")

# --- Settings loaded from config.yaml at runtime (see _load_settings) ---
JIRA_BASE_URL: str | None = None
JIRA_PAT_TOKEN: str | None = None
CONFLUENCE_URL: str | None = None
CONFLUENCE_PAT: str | None = None
CONFLUENCE_PAGE_ID: str | None = None
CONFLUENCE_SPACE_KEY: str | None = None

TAG_FIELD = "customfield_13905"

JIRA_JQL = (
    'project != "Issue Template" '
    'AND "Product Type" = "SMT PCBA" '
    'AND "NPI Location" = "Singapore" '
    'ORDER BY created ASC'
)

# --- EDM Oracle DB ---
EDM_HOST = "sgp01.sg.pepperl-fuchs.com"
EDM_PORT = 1521
EDM_SERVICE = "SGP01EDMEWA.WORLD"

# --- Excel ---
EXCEL_FILE = ROOT / "outputs" / "Pilot_DMR_Report.xlsx"
MAIN_TAB = "Active MR"
COMPLETED_TAB = "COMPLETED MR"

# --- Patterns ---
PE_PATTERN = re.compile(r'QD\s*-\s*\d+', re.IGNORECASE)
TE_PATTERN = re.compile(r'906\s*-\s*[A-Za-z0-9]+', re.IGNORECASE)
PT_PATTERN = re.compile(r'PT[A-Z0-9]{2}-[A-Z0-9]{4,5}', re.IGNORECASE)
MR_WEEK_PATTERN = re.compile(r'^\s*MR\s*Week\s*(\d+)\s*$', re.IGNORECASE)

# Container-key + ticked-checkbox patterns for the Status round-trip
BROWSE_KEY_PATTERN = re.compile(r'/browse/([A-Za-z][A-Za-z0-9]*-\d+)')
TASK_COMPLETE_PATTERN = re.compile(
    r'<ac:task-status>\s*complete\s*</ac:task-status>', re.IGNORECASE
)
TR_PATTERN = re.compile(r'<tr\b.*?</tr>', re.DOTALL | re.IGNORECASE)

# Column order (16 columns)
COL_KEYS = [
    "Container", "Type", "PT_Number", "PRSG_Number", "PRSG_Status",
    "SMT_Closure", "Doc_Closure", "Close_Date", "PE_Reports", "TE_Reports",
    "Handover_PE", "Handover_TE", "Ageing", "MR_Status", "Remarks",
    "Completion_Date"
]

HEADERS = [
    "Container Numbers", "Type (Pilot/DMR)", "PT Number", "PRSG Number",
    "PRSG Status", "SMT Build Closure Date", "Doc Closure Date",
    "Container Close Date", "PE Reports", "TE Reports",
    "Handover PE", "Handover TE", "Ageing (Days)", "MR Status",
    "Remarks", "Completion Date"
]

COL = {k: i + 1 for i, k in enumerate(COL_KEYS)}
NUM_COLS = len(HEADERS)

# Manual column indices (0-based for Confluence table parsing)
MANUAL_IDX = {
    "Handover_PE": 10, "Handover_TE": 11,
    "MR_Status": 13, "Remarks": 14,
}


def _load_settings(mode: str) -> None:
    """Populate the module-level connection globals from config.yaml."""
    global JIRA_BASE_URL, JIRA_PAT_TOKEN
    global CONFLUENCE_URL, CONFLUENCE_PAT, CONFLUENCE_PAGE_ID, CONFLUENCE_SPACE_KEY
    cfg = load_config(mode)
    JIRA_BASE_URL = cfg.jira_base_url
    JIRA_PAT_TOKEN = cfg.jira_pat
    CONFLUENCE_URL = cfg.confluence_base_url
    CONFLUENCE_PAT = cfg.confluence_pat
    CONFLUENCE_SPACE_KEY = cfg.confluence_space_key
    CONFLUENCE_PAGE_ID = str(cfg.pages.get("mr_status_report") or 560866215)


# =====================================================================
# JIRA API HELPERS
# =====================================================================
def make_session():
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {JIRA_PAT_TOKEN}", "Accept": "application/json"})
    s.verify = False
    return s


def fetch_children(session, parent_key):
    jql = f"issue in relation('{parent_key}', 'Project Children', Tasks, Deviations, level4)"
    params = {"jql": jql, "fields": "summary,status,resolutiondate", "maxResults": 50}
    try:
        resp = session.get(f"{JIRA_BASE_URL}/rest/api/2/search", params=params, timeout=15)
        if resp.status_code == 200:
            return [i for i in resp.json().get("issues", []) if i["key"] != parent_key]
    except Exception as e:
        print(f"  ⚠ Children fetch failed for {parent_key}: {e}")
    return []


def classify_container(tag_field_data, children, key=""):
    if not tag_field_data:
        log.debug(f"[{key}] SKIP: No tag field data")
        return None
    cv = str(tag_field_data.get("value", "") if isinstance(tag_field_data, dict) else tag_field_data).lower()
    cs = [c.get("fields", {}).get("summary", "").strip() for c in children]
    log.debug(f"[{key}] Category value: '{cv}' | Children: {cs}")

    if "pilot run" in cv and any(s == "QM P+L" for s in cs):
        log.info(f"[{key}] Classified: Pilot Run")
        return "Pilot Run"
    if ("direct manufacturing release" in cv or cv.startswith("dmr")) and any(s == "Direct Manufacturing Release" for s in cs):
        if len(children) > 1:
            extras = [s for s in cs if s != "Direct Manufacturing Release"]
            log.warning(f"[{key}] DMR has extra children beyond 'Direct Manufacturing Release': {extras}")
        log.info(f"[{key}] Classified: DMR Request")
        return "DMR Request"

    log.debug(f"[{key}] SKIP: No classification match (tag='{cv}', children={cs})")
    return None


def get_child_resolution_date(children, target):
    for c in children:
        if c.get("fields", {}).get("summary", "").strip().lower() == target.lower():
            rd = c.get("fields", {}).get("resolutiondate")
            if rd:
                return rd[:10]
    return ""


def fetch_comments(session, issue_key):
    pe, te = set(), set()
    try:
        resp = session.get(f"{JIRA_BASE_URL}/rest/api/2/issue/{issue_key}/comment", timeout=10)
        if resp.status_code == 200:
            for c in resp.json().get("comments", []):
                body = c.get("body", "")
                pe.update(PE_PATTERN.findall(body))
                te.update(TE_PATTERN.findall(body))
    except Exception:
        pass
    clean = lambda s: re.sub(r'\s+', '', s).upper()
    return ", ".join(sorted(clean(p) for p in pe)), ", ".join(sorted(clean(t) for t in te))


def extract_pt_number(summary):
    m = PT_PATTERN.search(summary or "")
    return m.group(0).upper() if m else ""


def process_issue(issue, session):
    key = issue["key"]
    fields = issue.get("fields", {})
    tag_data = fields.get(TAG_FIELD)
    if not tag_data or not isinstance(tag_data, dict):
        return None
    cv = tag_data.get("value", "").lower()
    if "pilot run" not in cv and "direct manufacturing release" not in cv and not cv.startswith("dmr"):
        return None

    # Filter: skip Won't Do resolutions
    resolution = fields.get("resolution")
    if resolution and isinstance(resolution, dict):
        res_name = resolution.get("name", "")
        if res_name.lower().replace("'", "’") in ("won't do", "won’t do", "wont do"):
            log.info(f"[{key}] SKIPPED: Resolution = '{res_name}' (Won't Do)")
            return None

    log.info(f"[{key}] Processing... (tag: {tag_data.get('value', '')})")
    children = fetch_children(session, key)
    category = classify_container(tag_data, children, key)
    if not category:
        return None

    summary = fields.get("summary", "")
    pt = extract_pt_number(summary)
    smt = get_child_resolution_date(children, "SMT Build")
    doc = get_child_resolution_date(children, "Documentation")
    pe, te = fetch_comments(session, key)
    created = fields.get("created", "")[:10]
    now = datetime.now()

    # Container close date (parent's resolution date)
    close_date = ""
    res_date = fields.get("resolutiondate")
    if res_date:
        close_date = res_date[:10]

    if category == "Pilot Run" and smt:
        ageing = (now - datetime.strptime(smt, "%Y-%m-%d")).days
    elif category == "Pilot Run":
        ageing = ""
    else:
        ageing = (now - datetime.strptime(created, "%Y-%m-%d")).days

    log.info(f"[{key}] {category} | PT={pt} | SMT={smt} | Doc={doc} | Close={close_date} | Ageing={ageing} | PE={pe} | TE={te}")

    return key, {
        "Container": key, "Type": category, "PT_Number": pt,
        "SMT_Closure": smt, "Doc_Closure": doc, "Close_Date": close_date,
        "PE_Reports": pe, "TE_Reports": te, "Ageing": ageing,
    }


def fetch_all_from_jira(session, ignore_keys):
    url = f"{JIRA_BASE_URL}/rest/api/2/search"
    all_issues, start_at = [], 0
    while True:
        resp = session.post(url, json={
            "jql": JIRA_JQL, "fields": ["summary", TAG_FIELD, "status", "created", "resolution", "resolutiondate"],
            "startAt": start_at, "maxResults": 100
        }, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        all_issues.extend(data.get("issues", []))
        total = data.get("total", 0)
        print(f"  Fetched {len(all_issues)}/{total} issues...")
        if len(all_issues) >= total:
            break
        start_at += 100

    print(f"\n📦 Total: {len(all_issues)} | Skipping: {len(ignore_keys)} completed")
    to_process = [i for i in all_issues if i["key"] not in ignore_keys]
    print(f"🔍 Processing {len(to_process)} containers...\n")

    results = {}
    with ThreadPoolExecutor(max_workers=10) as ex:
        futs = {ex.submit(process_issue, i, session): i["key"] for i in to_process}
        for f in as_completed(futs):
            try:
                r = f.result()
                if r:
                    results[r[0]] = r[1]
            except Exception as e:
                print(f"  ⚠ Error: {futs[f]}: {e}")

    print(f"\n✅ Found {len(results)} Pilot Run / DMR containers")
    return results


# =====================================================================
# EDM ORACLE DB
# =====================================================================
def edm_connect():
    try:
        import oracledb
    except ImportError:
        print("  ⚠ oracledb not installed — skipping EDM")
        return None
    try:
        oracledb.init_oracle_client()
    except Exception:
        pass
    try:
        dsn = oracledb.makedsn(EDM_HOST, EDM_PORT, service_name=EDM_SERVICE)
        return oracledb.connect(dsn=dsn, externalauth=True)
    except Exception as e:
        print(f"  ⚠ EDM connection failed: {e}")
        return None


def edm_lookup_prsg(conn, pt_numbers):
    if not conn or not pt_numbers:
        log.info("EDM: Skipped (no connection or no PT numbers)")
        return {}
    cur = conn.cursor()
    results = {}
    unique_pts = list(set(pt for pt in pt_numbers if pt))
    if not unique_pts:
        return {}
    log.info(f"EDM: Looking up {len(unique_pts)} PT numbers...")
    for i in range(0, len(unique_pts), 500):
        batch = unique_pts[i:i + 500]
        ph = ",".join(f":{j+1}" for j in range(len(batch)))
        try:
            cur.execute(f"""
                SELECT r.REF, r.DOCNUMBER, d.RELEASESTATE
                FROM ADMEDP.EDM_REFERENCES r
                JOIN ADMEDP.EDM_DOCS d ON d.DOCNUMBER = r.DOCNUMBER
                WHERE r.REF IN ({ph}) AND r.DOCNUMBER LIKE 'PRSG-%'
            """, batch)
            for pt, prsg, rs in cur.fetchall():
                pu = pt.upper()
                status = "Released" if rs == 9 else "Not Released"
                log.info(f"  EDM: PT={pu} → PRSG={prsg} | RELEASESTATE={rs} → {status}")
                if pu in results and results[pu]["status"] == "Released":
                    log.debug(f"  EDM: Skipping {prsg} for {pu} (already have a Released PRSG)")
                    continue
                results[pu] = {"prsg": prsg, "status": status}
        except Exception as e:
            print(f"  ⚠ EDM query error: {e}")
    found = sum(1 for r in results.values() if r["prsg"])
    released = sum(1 for r in results.values() if r["status"] == "Released")
    print(f"  📋 {found} PRSG links ({released} released)")
    cur.close()
    return results


# =====================================================================
# CONFLUENCE
# =====================================================================
def conf_session():
    s = requests.Session()
    s.headers.update({
        "Authorization": f"Bearer {CONFLUENCE_PAT}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    })
    s.verify = False
    return s


def parse_ticked_containers(html):
    """Scan raw storage HTML for rows whose Status checkbox is ticked.

    Independent of column position: for each <tr> that contains a completed
    task checkbox, grab the container key from its /browse/<KEY> link. Matching
    on the exact <ac:task-status>complete</ac:task-status> tag avoids the
    'incomplete' substring trap.
    """
    ticked = set()
    if not html:
        return ticked
    for chunk in TR_PATTERN.findall(html):
        if not TASK_COMPLETE_PATTERN.search(chunk):
            continue
        m = BROWSE_KEY_PATTERN.search(chunk)
        if m:
            ticked.add(m.group(1).strip())
    return ticked


def conf_read_manual_fields(csess):
    """Read Confluence page, parse tables, return manual fields + completed keys
    + completed rows + ticked-done container keys + version."""
    url = f"{CONFLUENCE_URL}/rest/api/content/{CONFLUENCE_PAGE_ID}?expand=body.storage,version"
    resp = csess.get(url, timeout=15)
    if resp.status_code != 200:
        log.warning(f"Confluence read failed: {resp.status_code}")
        return {}, set(), [], set(), 0

    page = resp.json()
    version = page["version"]["number"]
    html = page.get("body", {}).get("storage", {}).get("value", "")
    manual = {}
    completed = set()

    # NEW: which containers have a ticked Status box (parsed from raw html,
    # before macro/tag stripping mangles the task tags).
    ticked_done = parse_ticked_containers(html)
    if ticked_done:
        log.info(f"  Status checkbox ticked for: {sorted(ticked_done)}")

    if not html or "<table" not in html:
        return manual, completed, [], ticked_done, version

    # Smart macro handling: extract title text from status macros instead of deleting
    # Turns <ac:structured-macro ac:name="status">...<ac:parameter ac:name="title">DONE</ac:parameter>...</ac:structured-macro>
    # into just "DONE"
    def replace_status_macro(match):
        title_match = re.search(r'ac:name="title"[^>]*>([^<]*)<', match.group(0))
        return title_match.group(1) if title_match else ""

    clean_html = re.sub(
        r'<ac:structured-macro[^>]*ac:name="status"[^>]*>.*?</ac:structured-macro>',
        replace_status_macro, html, flags=re.DOTALL
    )
    # Strip any other non-status ac: macros and remaining ac: tags
    clean_html = re.sub(r'<ac:structured-macro[^>]*>.*?</ac:structured-macro>', '', clean_html, flags=re.DOTALL)
    clean_html = re.sub(r'</?ac:[^>]*>', '', clean_html)

    from html.parser import HTMLParser

    class TParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.in_td = False
            self.in_row = False
            self.in_table = 0
            self.cell_text = ""
            self.row_cells = []
            self.rows = []

        def handle_starttag(self, tag, attrs):
            if tag == "table":
                self.in_table += 1
                self.rows = []
            elif tag == "tr" and self.in_table:
                self.in_row = True
                self.row_cells = []
            elif tag in ("td", "th") and self.in_row:
                self.in_td = True
                self.cell_text = ""

        def handle_endtag(self, tag):
            if tag == "table":
                self.in_table -= 1
            elif tag == "tr" and self.in_row:
                self.in_row = False
                self.rows.append(self.row_cells)
            elif tag in ("td", "th") and self.in_td:
                self.in_td = False
                self.row_cells.append(self.cell_text.strip())

        def handle_data(self, data):
            if self.in_td:
                self.cell_text += data

    # Split HTML: skip MR Week table, parse Active MR and Completed MR
    # Structure: [MR Week section] <h2>Active MR</h2> [active table] <h2>COMPLETED MR</h2> [completed table]
    active_split = re.split(r'<h2[^>]*>.*?Active MR.*?</h2>', clean_html, flags=re.IGNORECASE | re.DOTALL)
    active_html = active_split[-1] if len(active_split) > 1 else clean_html

    parts = re.split(r'<h2[^>]*>.*?COMPLETED MR.*?</h2>', active_html, flags=re.IGNORECASE | re.DOTALL)

    # Parse active table (first part after Active MR heading)
    if parts:
        p = TParser()
        p.feed(parts[0])
        if p.rows:
            log.debug(f"  Active table: {len(p.rows)} rows, first row has {len(p.rows[0])} cells")
            if len(p.rows) > 1:
                log.debug(f"  Sample data row: {len(p.rows[1])} cells → {p.rows[1][:3]}...{p.rows[1][-3:] if len(p.rows[1]) > 3 else ''}")
        for row in p.rows[1:]:  # skip header
            if row and len(row) >= 10 and row[0]:
                key = row[0].strip()
                num_cells = len(row)
                m = {
                    "Handover_PE": row[MANUAL_IDX["Handover_PE"]] if num_cells > MANUAL_IDX["Handover_PE"] else "",
                    "Handover_TE": row[MANUAL_IDX["Handover_TE"]] if num_cells > MANUAL_IDX["Handover_TE"] else "",
                    "MR_Status": row[MANUAL_IDX["MR_Status"]] if num_cells > MANUAL_IDX["MR_Status"] else "WAITING",
                    "Remarks": row[MANUAL_IDX["Remarks"]] if num_cells > MANUAL_IDX["Remarks"] else "",
                }
                manual[key] = m
                # Log non-empty manual fields
                non_empty = {k: v for k, v in m.items() if v}
                if non_empty:
                    log.info(f"  [{key}] Manual fields preserved: {non_empty}")
                else:
                    log.debug(f"  [{key}] Parsed ({num_cells} cells) — no manual edits")

    # Parse completed table (second part) — preserve full row data
    completed_rows = []
    if len(parts) > 1:
        p2 = TParser()
        p2.feed(parts[1])
        log.debug(f"  Completed table: {len(p2.rows)} rows")
        for row in p2.rows[1:]:
            if row and row[0]:
                key = row[0].strip()
                completed.add(key)
                n = len(row)

                # Detect old (15-col, no Close_Date) vs new (16-col) layout
                if n >= 16:
                    # New layout: Close_Date at position 7
                    cr = {
                        "Container": key,
                        "Type": row[1], "PT_Number": row[2],
                        "PRSG_Number": row[3], "PRSG_Status": row[4],
                        "SMT_Closure": row[5], "Doc_Closure": row[6],
                        "Close_Date": row[7],
                        "PE_Reports": row[8], "TE_Reports": row[9],
                        "Handover_PE": row[10], "Handover_TE": row[11],
                        "Ageing": row[12], "MR_Status": "DONE",
                        "Remarks": row[14], "Completion_Date": row[15],
                    }
                else:
                    # Old layout (15 cols): no Close_Date column
                    cr = {
                        "Container": key,
                        "Type": row[1] if n > 1 else "",
                        "PT_Number": row[2] if n > 2 else "",
                        "PRSG_Number": row[3] if n > 3 else "",
                        "PRSG_Status": row[4] if n > 4 else "",
                        "SMT_Closure": row[5] if n > 5 else "",
                        "Doc_Closure": row[6] if n > 6 else "",
                        "Close_Date": "",
                        "PE_Reports": row[7] if n > 7 else "",
                        "TE_Reports": row[8] if n > 8 else "",
                        "Handover_PE": row[9] if n > 9 else "",
                        "Handover_TE": row[10] if n > 10 else "",
                        "Ageing": row[11] if n > 11 else "",
                        "MR_Status": "DONE",
                        "Remarks": row[13] if n > 13 else "",
                        "Completion_Date": row[14] if n > 14 else "",
                    }
                log.debug(f"  [{key}] Completed row parsed ({n} cells): Ageing={cr['Ageing']}, Remarks={cr['Remarks']}, Date={cr['Completion_Date']}")
                completed_rows.append(cr)

    return manual, completed, completed_rows, ticked_done, version


def build_html(active_rows, completed_rows):
    """Build Confluence storage format HTML with MR Week priority table."""
    from html import escape as html_escape
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Unique task-id source for the Status checkboxes (ids must be unique per page).
    task_seq = [0]

    def checkbox_cell():
        task_seq[0] += 1
        return (f'<td {cen}><ac:task-list>'
                f'<ac:task><ac:task-id>{task_seq[0]}</ac:task-id>'
                f'<ac:task-status>incomplete</ac:task-status>'
                f'<ac:task-body></ac:task-body></ac:task>'
                f'</ac:task-list></td>\n')

    def esc(v):
        """Escape HTML special chars in user data (&, <, >, quotes)."""
        if v is None or v == "":
            return ""
        return html_escape(str(v), quote=True)

    def status_macro(text, colour):
        return (f'<ac:structured-macro ac:name="status">'
                f'<ac:parameter ac:name="colour">{colour}</ac:parameter>'
                f'<ac:parameter ac:name="title">{esc(text)}</ac:parameter>'
                f'</ac:structured-macro>')

    def prsg_badge(s):
        if s == "Released": return status_macro("Released", "Green")
        if s == "Not Released": return status_macro("Not Released", "Red")
        return ""

    def mr_badge(s):
        su = str(s).strip().upper()
        if su == "DONE": return status_macro("DONE", "Green")
        if su == "IN PROGRESS": return status_macro("IN PROGRESS", "Yellow")
        if su == "WAITING": return status_macro("WAITING", "Grey")
        return str(s)

    def ho_badge(s):
        """Render Handover cell: badge for standard values, plain text for free-form."""
        if not s:
            return ""
        su = str(s).strip().upper()
        if su == "DONE": return status_macro("DONE", "Green")
        if su == "IN PROGRESS": return status_macro("IN PROGRESS", "Yellow")
        # Free text — preserve as-is with escaping
        return esc(s)

    def age_td(v):
        if v == "" or v is None:
            return '<td style="text-align:center">-</td>'
        try:
            v = int(v)
        except Exception:
            return f'<td style="text-align:center">{v}</td>'
        if v >= 25:
            return f'<td style="background-color:#FF4444;color:white;font-weight:bold;text-align:center">{v}</td>'
        if v >= 20:
            return f'<td style="background-color:#FFA500;font-weight:bold;text-align:center">{v}</td>'
        if v >= 15:
            return f'<td style="background-color:#FFFF00;text-align:center">{v}</td>'
        return f'<td style="text-align:center">{v}</td>'

    def link(key):
        return f'<a href="{JIRA_BASE_URL}/browse/{key}">{key}</a>'

    hdr_style = 'style="background-color:#4A90D9;color:white;text-align:center;padding:8px;font-weight:bold"'
    hdr_style2 = 'style="background-color:#5B7FA5;color:white;text-align:center;padding:8px;font-weight:bold"'
    hdr_style_mr = 'style="background-color:#8E44AD;color:white;text-align:center;padding:8px;font-weight:bold"'
    cen = 'style="text-align:center"'

    # Active table headers (exclude Completion Date); a "Status" tick-box column
    # is appended at the end (handled when rendering each table).
    active_headers = HEADERS[:-1]

    def render_active_row(r):
        """Render one active row (shared between MR Week table and Active table)."""
        tc = "#D6EAF8" if r["Type"] == "Pilot Run" else "#D5F5E3" if r["Type"] == "DMR Request" else "white"
        h = '<tr>\n'
        h += f'<td {cen}>{link(r["Container"])}</td>\n'
        h += f'<td style="background-color:{tc};text-align:center">{esc(r["Type"])}</td>\n'
        h += f'<td {cen}>{esc(r.get("PT_Number",""))}</td>\n'
        h += f'<td {cen}>{esc(r.get("PRSG_Number",""))}</td>\n'
        h += f'<td {cen}>{prsg_badge(r.get("PRSG_Status",""))}</td>\n'
        h += f'<td {cen}>{esc(r.get("SMT_Closure",""))}</td>\n'
        h += f'<td {cen}>{esc(r.get("Doc_Closure",""))}</td>\n'
        h += f'<td {cen}>{esc(r.get("Close_Date",""))}</td>\n'
        h += f'<td>{esc(r.get("PE_Reports",""))}</td>\n'
        h += f'<td>{esc(r.get("TE_Reports",""))}</td>\n'
        h += f'<td {cen}>{ho_badge(r.get("Handover_PE",""))}</td>\n'
        h += f'<td {cen}>{ho_badge(r.get("Handover_TE",""))}</td>\n'
        h += age_td(r.get("Ageing", "")) + '\n'
        h += f'<td {cen}>{mr_badge(r.get("MR_Status",""))}</td>\n'
        h += f'<td>{esc(r.get("Remarks",""))}</td>\n'
        h += checkbox_cell()
        h += '</tr>\n'
        return h

    # --- MR Week Priority Table ---
    mr_week_rows = []
    for r in active_rows:
        remarks = str(r.get("Remarks", ""))
        match = MR_WEEK_PATTERN.search(remarks)
        if match:
            week_num = int(match.group(1))
            mr_week_rows.append((week_num, r))

    mr_week_rows.sort(key=lambda x: x[0])

    # Info bar
    h = f'<p><strong>Last Updated:</strong> {now} &nbsp;|&nbsp; '
    h += f'<strong>Active:</strong> {len(active_rows)} &nbsp;|&nbsp; '
    h += f'<strong>MR Week Tagged:</strong> {len(mr_week_rows)} &nbsp;|&nbsp; '
    h += f'<strong>Completed:</strong> {len(completed_rows)}</p>\n'

    # MR Week table (only shown if there are tagged rows)
    if mr_week_rows:
        h += '<h2>MR Week Schedule</h2>\n'
        h += '<p><em>Containers tagged with "MR Week XX" in Remarks, sorted by week number. '
        h += 'Automatically removed when MR Status is DONE.</em></p>\n'
        h += '<table><thead><tr>\n'
        h += f'<th {hdr_style_mr}>MR Week</th>\n'
        for hd in active_headers:
            h += f'<th {hdr_style_mr}>{hd}</th>\n'
        h += f'<th {hdr_style_mr}>Status</th>\n'
        h += '</tr></thead>\n<tbody>\n'
        for week_num, r in mr_week_rows:
            h += f'<tr>\n<td style="text-align:center;font-weight:bold;background-color:#F5EEF8">Week {week_num}</td>\n'
            # Render the row without the outer <tr> tags
            tc = "#D6EAF8" if r["Type"] == "Pilot Run" else "#D5F5E3" if r["Type"] == "DMR Request" else "white"
            h += f'<td {cen}>{link(r["Container"])}</td>\n'
            h += f'<td style="background-color:{tc};text-align:center">{esc(r["Type"])}</td>\n'
            h += f'<td {cen}>{esc(r.get("PT_Number",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("PRSG_Number",""))}</td>\n'
            h += f'<td {cen}>{prsg_badge(r.get("PRSG_Status",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("SMT_Closure",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("Doc_Closure",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("Close_Date",""))}</td>\n'
            h += f'<td>{esc(r.get("PE_Reports",""))}</td>\n'
            h += f'<td>{esc(r.get("TE_Reports",""))}</td>\n'
            h += f'<td {cen}>{ho_badge(r.get("Handover_PE",""))}</td>\n'
            h += f'<td {cen}>{ho_badge(r.get("Handover_TE",""))}</td>\n'
            h += age_td(r.get("Ageing", "")) + '\n'
            h += f'<td {cen}>{mr_badge(r.get("MR_Status",""))}</td>\n'
            h += f'<td>{esc(r.get("Remarks",""))}</td>\n'
            h += checkbox_cell()
            h += '</tr>\n'
        h += '</tbody></table>\n'

    # --- Active MR Table ---
    h += '<h2>Active MR</h2>\n<table><thead><tr>\n'
    for hd in active_headers:
        h += f'<th {hdr_style}>{hd}</th>\n'
    h += f'<th {hdr_style}>Status</th>\n'
    h += '</tr></thead>\n<tbody>\n'
    for r in active_rows:
        h += render_active_row(r)
    h += '</tbody></table>\n'

    # --- Completed MR Table ---
    if completed_rows:
        h += '<h2>COMPLETED MR</h2>\n<table><thead><tr>\n'
        for hd in HEADERS:
            h += f'<th {hdr_style2}>{hd}</th>\n'
        h += '</tr></thead>\n<tbody>\n'
        for r in completed_rows:
            h += '<tr>\n'
            h += f'<td {cen}>{link(r["Container"])}</td>\n'
            h += f'<td {cen}>{esc(r["Type"])}</td>\n'
            h += f'<td {cen}>{esc(r.get("PT_Number",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("PRSG_Number",""))}</td>\n'
            h += f'<td {cen}>{prsg_badge(r.get("PRSG_Status",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("SMT_Closure",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("Doc_Closure",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("Close_Date",""))}</td>\n'
            h += f'<td>{esc(r.get("PE_Reports",""))}</td>\n'
            h += f'<td>{esc(r.get("TE_Reports",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("Handover_PE",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("Handover_TE",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("Ageing",""))}</td>\n'
            h += f'<td {cen}>{status_macro("DONE","Green")}</td>\n'
            h += f'<td>{esc(r.get("Remarks",""))}</td>\n'
            h += f'<td {cen}>{esc(r.get("Completion_Date",""))}</td>\n'
            h += '</tr>\n'
        h += '</tbody></table>\n'

    return h


def conf_update(csess, html, version, retry=True, dry_run=False):
    if dry_run:
        log.info(f"DRY-RUN: would publish v{version + 1} ({len(html)} bytes) to page {CONFLUENCE_PAGE_ID} — NOT sent.")
        print(f"  🧪 DRY-RUN: page NOT updated (would be v{version + 1}, {len(html)} bytes)")
        return True
    payload = {
        "id": CONFLUENCE_PAGE_ID,
        "type": "page",
        "title": "Express Ops MR Status",
        "space": {"key": CONFLUENCE_SPACE_KEY},
        "version": {"number": version + 1},
        "body": {"storage": {"value": html, "representation": "storage"}}
    }
    resp = csess.put(f"{CONFLUENCE_URL}/rest/api/content/{CONFLUENCE_PAGE_ID}", json=payload, timeout=30)
    if resp.status_code == 200:
        log.info(f"Confluence updated (v{version + 1})")
        return True

    # Log full error response
    log.error(f"Confluence update FAILED: HTTP {resp.status_code}")
    log.error(f"Response body: {resp.text[:1000]}")
    log.error(f"Sent version: {version + 1} | Page ID: {CONFLUENCE_PAGE_ID} | Space: {CONFLUENCE_SPACE_KEY}")

    # If 409 (conflict) or 400 with version issue, re-fetch and retry once
    if retry and resp.status_code in (400, 409):
        log.warning("Possible version conflict — re-fetching latest page version and retrying...")
        try:
            check = csess.get(f"{CONFLUENCE_URL}/rest/api/content/{CONFLUENCE_PAGE_ID}?expand=version", timeout=15)
            if check.status_code == 200:
                latest_version = check.json()["version"]["number"]
                log.info(f"Latest page version is {latest_version} (we sent {version + 1})")
                if latest_version != version:
                    log.info(f"Retrying with version {latest_version + 1}...")
                    return conf_update(csess, html, latest_version, retry=False, dry_run=dry_run)
        except Exception as e:
            log.error(f"Retry fetch failed: {e}")
    return False


# =====================================================================
# EXCEL (backup)
# =====================================================================
def write_excel(active_rows, completed_rows):
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.active.title = MAIN_TAB
    wb.create_sheet(COMPLETED_TAB)
    main_ws, comp_ws = wb[MAIN_TAB], wb[COMPLETED_TAB]

    fill, font, align, border = (
        PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid"),
        Font(bold=True, color="FFFFFF", name="Arial", size=10),
        Alignment(horizontal="center", vertical="center", wrap_text=True),
        Border(bottom=Side(style="thin", color="000000"))
    )
    for ws in [main_ws, comp_ws]:
        for i, hd in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=i, value=hd)
            c.fill, c.font, c.alignment, c.border = fill, font, align, border
        ws.row_dimensions[1].height = 30

    body_font = Font(name="Arial", size=10)
    cen = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
    left_cols = {COL["PE_Reports"], COL["TE_Reports"], COL["Remarks"]}

    for idx, row in enumerate(active_rows):
        r = idx + 2
        for k, cn in COL.items():
            cell = main_ws.cell(row=r, column=cn, value=row.get(k, ""))
            cell.font = body_font
            cell.alignment = lft if cn in left_cols else cen

        # Hyperlink
        lc = main_ws.cell(row=r, column=COL["Container"])
        lc.hyperlink = f"{JIRA_BASE_URL}/browse/{row['Container']}"
        lc.font = Font(name="Arial", size=10, color="0563C1", underline="single")

        # Ageing color
        ac = main_ws.cell(row=r, column=COL["Ageing"])
        v = ac.value
        if isinstance(v, (int, float)):
            if v >= 25:
                ac.fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
                ac.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            elif v >= 20:
                ac.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            elif v >= 15:
                ac.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Type color
        tc = main_ws.cell(row=r, column=COL["Type"])
        if tc.value == "Pilot Run":
            tc.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        elif tc.value == "DMR Request":
            tc.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

        # PRSG color
        pc = main_ws.cell(row=r, column=COL["PRSG_Status"])
        if pc.value == "Released":
            pc.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            pc.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        elif pc.value == "Not Released":
            pc.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            pc.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    for idx, row in enumerate(completed_rows):
        r = idx + 2
        for k, cn in COL.items():
            comp_ws.cell(row=r, column=cn, value=row.get(k, ""))

    # Dropdown only for MR Status (N) — Handover PE/TE allow free text
    dv = DataValidation(type="list", formula1='"WAITING,IN PROGRESS,DONE"', allow_blank=True)
    main_ws.add_data_validation(dv)
    dv.add(f"N2:N{max(main_ws.max_row + 50, 200)}")

    widths = {"A": 18, "B": 14, "C": 14, "D": 14, "E": 14, "F": 18, "G": 18,
              "H": 16, "I": 20, "J": 20, "K": 14, "L": 14, "M": 12, "N": 14, "O": 25, "P": 16}
    for ws in [main_ws, comp_ws]:
        for c, w in widths.items():
            ws.column_dimensions[c].width = w
        ws.freeze_panes = "A2"

    wb.save(EXCEL_FILE)


# =====================================================================
# MAIN
# =====================================================================
def run(dry_run=False):
    log.info("")
    log.info("=" * 70)
    log.info("  NEW RUN: Pilot Run & DMR - MR Tracking Report" + ("  [DRY-RUN]" if dry_run else ""))
    log.info(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 70)

    # 1. Read Confluence (source of truth for manual fields + completed history)
    log.info("Reading Confluence page...")
    csess = conf_session()
    try:
        manual, completed_keys, prev_completed_rows, ticked_done, page_ver = conf_read_manual_fields(csess)
        log.info(f"  Confluence: {len(manual)} active rows | {len(completed_keys)} completed | "
                 f"{len(ticked_done)} ticked | page v{page_ver}")
    except Exception as e:
        print(f"  ⚠ Confluence read error: {e}")
        manual, completed_keys, prev_completed_rows, ticked_done, page_ver = {}, set(), [], set(), 0

    # 2. Fetch Jira
    session = make_session()
    jira_data = fetch_all_from_jira(session, completed_keys)
    if not jira_data:
        print("\nNo active containers found.")
        # Still re-publish to preserve completed table
        if prev_completed_rows:
            print("\n📝 Re-publishing Confluence (preserving completed table)...")
            html = build_html([], prev_completed_rows)
            conf_update(csess, html, page_ver, dry_run=dry_run)
        return

    # 3. EDM lookup
    pts = [d["PT_Number"] for d in jira_data.values() if d.get("PT_Number")]
    edm_conn = edm_connect()
    prsg_map = edm_lookup_prsg(edm_conn, pts)
    if edm_conn:
        edm_conn.close()

    for k, d in jira_data.items():
        info = prsg_map.get(d.get("PT_Number", "").upper(), {})
        d["PRSG_Number"] = info.get("prsg", "")
        d["PRSG_Status"] = info.get("status", "")
        if d["PT_Number"]:
            log.info(f"[{k}] PT={d['PT_Number']} → PRSG={d['PRSG_Number'] or 'NOT FOUND'} | Status={d['PRSG_Status'] or 'N/A'}")
        else:
            log.warning(f"[{k}] No PT number extracted from summary — PRSG lookup skipped")

    # 4. Merge with manual fields
    log.info("=" * 50)
    log.info("MERGE: Combining Jira/EDM data with manual fields")
    log.info("=" * 50)
    active, newly_done, auto_done, manual_ticked = [], [], 0, 0
    for k, d in jira_data.items():
        m = manual.get(k, {})
        row = {
            "Container": k, "Type": d["Type"],
            "PT_Number": d.get("PT_Number", ""),
            "PRSG_Number": d.get("PRSG_Number", ""),
            "PRSG_Status": d.get("PRSG_Status", ""),
            "SMT_Closure": d.get("SMT_Closure", ""),
            "Doc_Closure": d.get("Doc_Closure", ""),
            "Close_Date": d.get("Close_Date", ""),
            "PE_Reports": d.get("PE_Reports", ""),
            "TE_Reports": d.get("TE_Reports", ""),
            "Handover_PE": m.get("Handover_PE", ""),
            "Handover_TE": m.get("Handover_TE", ""),
            "Ageing": d.get("Ageing", ""),
            "MR_Status": m.get("MR_Status", "WAITING"),
            "Remarks": m.get("Remarks", ""),
            "Completion_Date": "",
        }

        # Log the PRSG check decision
        prsg_status = d.get("PRSG_Status", "")
        current_mr = row["MR_Status"]
        log.debug(f"[{k}] PRSG_Status='{prsg_status}' | Current MR_Status='{current_mr}' (from {'Confluence' if m else 'default'})")

        # Auto-DONE if PRSG released
        if prsg_status == "Released" and str(current_mr).strip().upper() != "DONE":
            row["MR_Status"] = "DONE"
            auto_done += 1
            log.info(f"[{k}] ✅ AUTO-DONE: PRSG {d['PRSG_Number']} is Released → MR Status changed to DONE")
        elif prsg_status == "Released" and str(current_mr).strip().upper() == "DONE":
            log.debug(f"[{k}] Already DONE (PRSG Released)")
        elif prsg_status == "Not Released":
            log.debug(f"[{k}] PRSG {d.get('PRSG_Number','')} Not Released → no auto-DONE")
        elif not prsg_status:
            log.debug(f"[{k}] No PRSG found → no auto-DONE")

        # NEW: manual settle — Status checkbox ticked on the page moves it to DONE,
        # even with no PRSG (projects that don't need to go for MR).
        if k in ticked_done and str(row["MR_Status"]).strip().upper() != "DONE":
            row["MR_Status"] = "DONE"
            manual_ticked += 1
            log.info(f"[{k}] ☑ TICKED-DONE: Status checkbox ticked → moving to COMPLETED MR")

        # Route to active or completed
        final_mr = str(row["MR_Status"]).strip().upper()
        if final_mr == "DONE":
            row["Completion_Date"] = datetime.now().strftime("%Y-%m-%d")
            newly_done.append(row)
            log.info(f"[{k}] → COMPLETED MR (MR_Status=DONE)")
        else:
            active.append(row)
            log.debug(f"[{k}] → Active (MR_Status={row['MR_Status']})")

    # Sort active by ageing desc
    active.sort(key=lambda r: (-r["Ageing"], r["Container"]) if isinstance(r.get("Ageing"), (int, float)) else (999, r["Container"]))

    # Combine: previously completed + newly completed
    all_completed = prev_completed_rows + newly_done

    print(f"\n📊 Active: {len(active)} | Newly Done: {len(newly_done)} | Previously Done: {len(prev_completed_rows)} "
          f"| Auto-DONE: {auto_done} | Ticked-DONE: {manual_ticked}")

    # 5. Publish Confluence
    print("\n📝 Publishing to Confluence..." + ("  (DRY-RUN)" if dry_run else ""))
    html = build_html(active, all_completed)
    conf_update(csess, html, page_ver, dry_run=dry_run)

    # 6. Excel backup
    print("\n💾 Saving Excel backup...")
    try:
        write_excel(active, all_completed)
        print(f"  ✅ {EXCEL_FILE}")
    except Exception as e:
        print(f"  ⚠ Excel failed: {e}")

    # Summary
    pr = sum(1 for r in active if r["Type"] == "Pilot Run")
    dm = sum(1 for r in active if r["Type"] == "DMR Request")
    print(f"\n{'='*60}")
    print(f"  SUMMARY{'  [DRY-RUN — page NOT updated]' if dry_run else ''}")
    print(f"  Pilot Runs    : {pr}")
    print(f"  DMR Requests  : {dm}")
    print(f"  Total Active  : {len(active)}")
    print(f"  Newly Done    : {len(newly_done)}")
    print(f"  Total Done    : {len(all_completed)}")
    print(f"  Auto-DONE     : {auto_done}")
    print(f"  Ticked-DONE   : {manual_ticked}")
    print(f"  📄 Confluence : {CONFLUENCE_URL}/spaces/{CONFLUENCE_SPACE_KEY}/pages/{CONFLUENCE_PAGE_ID}")
    print(f"  📊 Excel      : {EXCEL_FILE}")
    print(f"{'='*60}")


def main():
    ap = argparse.ArgumentParser(description="MR Status Report (Pilot Run & DMR)")
    g = ap.add_mutually_exclusive_group()
    g.add_argument("--mock", action="store_const", const="mock", dest="mode",
                   help="mock mode (no live systems) — this task is live-only, so it is a no-op")
    g.add_argument("--live", action="store_const", const="live", dest="mode",
                   help="connect to live JIRA / EDM / Confluence")
    ap.add_argument("--dry-run", action="store_true",
                    help="read live data and build the page, but do NOT publish to Confluence")
    ap.set_defaults(mode="mock")
    args = ap.parse_args()

    if args.mode == "mock":
        print("MR Status Report is a LIVE-ONLY task (needs JIRA + EDM + Confluence, "
              "and EDMAdmin.exe for the PRSG lookup).")
        print("Use:  EDMAdmin.exe -m tasks.mr_status_report.main --live")
        print("Preview without publishing:  ... --live --dry-run")
        return

    _load_settings("live")
    run(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
